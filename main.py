import os 
from openpyxl import load_workbook
from PyPDF2 import PdfFileReader, PdfFileWriter
from pdfminer.high_level import extract_text
   
def odczytaj_nrb_z_excela(sciezka_do_pliku):
    wb = load_workbook(sciezka_do_pliku)
    sheet = wb.active

    if sheet.max_row < 3:
        print("Arkusz Excela jest pusty lub zawiera za mało wierszy.")
        return []

    nrb_list = []
    for row in sheet.iter_rows(min_row=3, min_col=4, max_col=4):
        nrb = row[0].value
        if nrb:
            nrb_czysty = nrb.replace(' ', '')
            nrb_list.append(nrb_czysty)
            print(f'Odczytano NRB: {nrb_czysty}')
    return nrb_list

def extract_text_from_pdf_page(pdf_path, page_number):
    with open(pdf_path, 'rb') as f_in:
        return extract_text(f_in, page_numbers=[page_number])

def usun_klienta_z_pdf(sciezka_do_pdf, nrb_list):
    reader = PdfFileReader(sciezka_do_pdf)
    writer = PdfFileWriter()

    removing_client = False
    for i in range(reader.numPages):
        page_text = extract_text_from_pdf_page(sciezka_do_pdf, i)
       
        for nrb in nrb_list:
            if nrb in page_text.replace(' ', ''):
                removing_client = True
                break
       
        if "Termin płatności:" in page_text and removing_client:
            removing_client = False
            continue  

        if not removing_client:
            writer.addPage(reader.getPage(i))

    temp_pdf_path = sciezka_do_pdf.replace('.pdf', '_temp.pdf')
    with open(temp_pdf_path, 'wb') as f_out:
        writer.write(f_out)

    os.remove(sciezka_do_pdf)
    os.rename(temp_pdf_path, sciezka_do_pdf)


def przetworz_pdf():
    if not os.path.exists(excel_file_path):
        print(f"Plik Excela {excel_file_path} nie istnieje.")
        return

    if not os.path.exists(pdf_folder_path) or not os.path.isdir(pdf_folder_path):
        print(f"Folder {pdf_folder_path} nie istnieje.")
        return

    nrb_list = odczytaj_nrb_z_excela(excel_file_path)
    if not nrb_list:
        print("Brak NRB do przetworzenia.")
        return

    pdf_files = [os.path.join(pdf_folder_path, f) for f in os.listdir(pdf_folder_path) if f.endswith('.pdf')]

    for pdf_file in pdf_files:
        usun_klienta_z_pdf(pdf_file, nrb_list)

excel_file_path = '.xlsx'
pdf_folder_path = ''

przetworz_pdf()